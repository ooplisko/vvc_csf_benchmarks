from __future__ import annotations

import argparse
import csv
import html
import logging
import math
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from metrics.bd_rate import bd_psnr, bd_rate
from metrics.registry import METRICS, METRIC_CHART_LABELS

logger = logging.getLogger(__name__)


# ====================================================================================================================
# CSV summaries
# ====================================================================================================================


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig") as stream:
        return list(csv.DictReader(stream))


def f(row: dict[str, str], key: str) -> float:
    try:
        return float(row.get(key, "0") or 0)
    except ValueError:
        return 0.0


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, all_metrics: list[dict[str, str]], same_qp: list[dict[str, object]], bd_summary: list[dict[str, object]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError("XLSX output requires openpyxl. Install it with: pip install -r requirements.txt") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_sheet(workbook, "All metrics", all_metrics)
    _add_sheet(workbook, "Same-QP summary", same_qp)
    _add_sheet(workbook, "BD-Rate", bd_summary)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for sheet in workbook.worksheets:
        _style_sheet(sheet)
        _color_delta_columns(sheet, green, red)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def same_qp_deltas(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    by_key: dict[tuple[str, int], dict[str, dict[str, str]]] = defaultdict(dict)
    for row in rows:
        by_key[(row["image"], int(row["qp"]))][row["mode"]] = row

    output = []
    for (image, qp), modes in sorted(by_key.items()):
        if "baseline" not in modes or "csf" not in modes:
            continue
        base = modes["baseline"]
        csf = modes["csf"]
        out: dict[str, object] = {
            "image": image,
            "qp": qp,
            "baseline_bpp": f(base, "bpp"),
            "csf_bpp": f(csf, "bpp"),
            "bpp_delta_pct": _pct(f(csf, "bpp"), f(base, "bpp")),
            "compression_ratio_delta_pct": _pct(f(csf, "compression_ratio"), f(base, "compression_ratio")),
        }
        for metric in METRICS:
            out[f"{metric}_delta"] = f(csf, metric) - f(base, metric)
        output.append(out)
    return output


def equal_bpp_summary(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    by_image: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        by_image[row["image"]][row["mode"]].append(row)

    output = []
    for image, modes in sorted(by_image.items()):
        baseline = sorted(modes["baseline"], key=lambda row: f(row, "bpp"))
        csf = sorted(modes["csf"], key=lambda row: f(row, "bpp"))
        if len(baseline) < 2 or len(csf) < 2:
            continue

        low = max(f(csf[0], "bpp"), f(baseline[0], "bpp"))
        high = min(f(csf[-1], "bpp"), f(baseline[-1], "bpp"))
        samples = [row for row in baseline if low <= f(row, "bpp") <= high]
        out: dict[str, object] = {"image": image, "sample_count": len(samples)}
        for metric in METRICS:
            deltas = []
            for base_row in samples:
                csf_metric = _interp_metric(csf, f(base_row, "bpp"), metric)
                if csf_metric is not None:
                    deltas.append(csf_metric - f(base_row, metric))
            out[f"{metric}_equal_bpp_delta"] = sum(deltas) / len(deltas) if deltas else 0.0
        output.append(out)
    return output


def bd_rate_summary(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    by_image: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        by_image[row["image"]][row["mode"]].append(row)

    output = []
    for image, modes in sorted(by_image.items()):
        if "baseline" not in modes or "csf" not in modes:
            continue
        baseline = sorted(modes["baseline"], key=lambda row: f(row, "bpp"))
        csf = sorted(modes["csf"], key=lambda row: f(row, "bpp"))
        if len(baseline) < 2 or len(csf) < 2:
            continue
        for metric in METRICS:
            output.append(
                {
                    "image": image,
                    "metric": metric,
                    "bd_rate_pct": _empty_if_none(
                        bd_rate(
                            [f(row, "bpp") for row in baseline],
                            [f(row, metric) for row in baseline],
                            [f(row, "bpp") for row in csf],
                            [f(row, metric) for row in csf],
                        )
                    ),
                    "bd_psnr_delta": _empty_if_none(
                        bd_psnr(
                            [f(row, "bpp") for row in baseline],
                            [f(row, metric) for row in baseline],
                            [f(row, "bpp") for row in csf],
                            [f(row, metric) for row in csf],
                        )
                    ),
                }
            )
    return output


def aggregate_bd_rate(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    summary = []
    by_metric: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        by_metric[str(row["metric"])].append(row)

    for metric in METRICS:
        items = by_metric.get(metric, [])
        bd_rates = _numeric_values(items, "bd_rate_pct")
        bd_psnrs = _numeric_values(items, "bd_psnr_delta")
        summary.append(
            {
                "metric": metric,
                "valid_images": len(bd_rates),
                "bd_rate_pct_mean": _mean_values(bd_rates),
                "bd_rate_pct_min": min(bd_rates) if bd_rates else "",
                "bd_rate_pct_max": max(bd_rates) if bd_rates else "",
                "bd_psnr_delta_mean": _mean_values(bd_psnrs),
                "bd_psnr_delta_min": min(bd_psnrs) if bd_psnrs else "",
                "bd_psnr_delta_max": max(bd_psnrs) if bd_psnrs else "",
            }
        )
    return summary


def per_image_summary(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    by_image: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        by_image[str(row["image"])].append(row)

    output = []
    for image, items in sorted(by_image.items()):
        out: dict[str, object] = {
            "image": image,
            "qp_points": len(items),
            "bpp_delta_pct_mean": _mean(items, "bpp_delta_pct"),
            "compression_ratio_delta_pct_mean": _mean(items, "compression_ratio_delta_pct"),
        }
        for metric in METRICS:
            out[f"{metric}_delta_mean"] = _mean(items, f"{metric}_delta")
        output.append(out)
    return output


def aggregate_summary(rows: list[dict[str, object]], output: Path) -> None:
    metric_keys = [key for key in rows[0] if key.endswith("_delta") or key.endswith("_equal_bpp_delta")]
    summary = []
    for key in metric_keys:
        values = [float(row[key]) for row in rows]
        summary.append(
            {
                "metric": key,
                "mean": sum(values) / len(values),
                "min": min(values),
                "max": max(values),
            }
        )
    write_csv(output, summary)


def _add_sheet(workbook: object, title: str, rows: list[dict[str, object]] | list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _style_sheet(sheet: object) -> None:
    if sheet.max_row == 0 or sheet.max_column == 0:
        return
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        width = max(len(header), 12)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 34)


def _color_delta_columns(sheet: object, green: object, red: object) -> None:
    from openpyxl.formatting.rule import CellIsRule

    if sheet.max_row < 2:
        return
    for cell in sheet[1]:
        header = str(cell.value or "").lower()
        if "delta" not in header and "bd_rate" not in header:
            continue
        column = cell.column_letter
        cell_range = f"{column}2:{column}{sheet.max_row}"
        sheet.conditional_formatting.add(cell_range, CellIsRule(operator="greaterThan", formula=["0"], fill=green))
        sheet.conditional_formatting.add(cell_range, CellIsRule(operator="lessThan", formula=["0"], fill=red))


def render_metric_charts(rows: list[dict[str, str]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for metric in METRICS:
        points: dict[str, list[dict[str, float | int]]] = {"baseline": [], "csf": []}
        grouped: dict[tuple[str, int], list[dict[str, str]]] = defaultdict(list)
        for row in rows:
            grouped[(row["mode"], int(row["qp"]))].append(row)
        for (mode, _qp), group in grouped.items():
            bpp = sum(f(row, "bpp") for row in group) / len(group)
            quality = sum(f(row, metric) for row in group) / len(group)
            points[mode].append({"bpp": bpp, "quality": quality, "qp": int(group[0]["qp"])})
        for mode in points:
            points[mode].sort(key=lambda point: float(point["bpp"]))
        (output_dir / f"rd_{metric}.svg").write_text(_svg_chart(metric, points), encoding="utf-8")


def render_per_image_qp_charts(rows: list[dict[str, str]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["image"]].append(row)

    for image, image_rows in sorted(grouped.items()):
        image_dir = output_dir / _safe_name(image)
        image_dir.mkdir(parents=True, exist_ok=True)
        for metric in METRICS:
            points: dict[str, list[dict[str, float | int]]] = {"baseline": [], "csf": []}
            for row in image_rows:
                mode = row["mode"]
                if mode in points:
                    points[mode].append({"qp": int(row["qp"]), "quality": f(row, metric)})
            for mode in points:
                points[mode].sort(key=lambda point: int(point["qp"]))
            (image_dir / f"qp_{metric}.svg").write_text(_svg_qp_chart(image, metric, points), encoding="utf-8")


# ====================================================================================================================
# Numeric helpers
# ====================================================================================================================


def _pct(new_value: float, old_value: float) -> float:
    return ((new_value - old_value) / old_value * 100.0) if old_value else 0.0


def _empty_if_none(value: float | None) -> float | str:
    return "" if value is None else value


def _mean(rows: list[dict[str, object]], key: str) -> float:
    values = [float(row[key]) for row in rows]
    return sum(values) / len(values) if values else 0.0


def _numeric_values(rows: list[dict[str, object]], key: str) -> list[float]:
    values = []
    for row in rows:
        value = row.get(key, "")
        if value == "":
            continue
        values.append(float(value))
    return values


def _mean_values(values: list[float]) -> float | str:
    return sum(values) / len(values) if values else ""


def _interp_metric(rows: list[dict[str, str]], bpp: float, metric: str) -> float | None:
    x = math.log(max(bpp, 1e-12))
    points = [(math.log(max(f(row, "bpp"), 1e-12)), f(row, metric)) for row in rows]
    for (x0, y0), (x1, y1) in zip(points, points[1:]):
        if min(x0, x1) <= x <= max(x0, x1):
            if x0 == x1:
                return y0
            ratio = (x - x0) / (x1 - x0)
            return y0 + ratio * (y1 - y0)
    return None


# ====================================================================================================================
# SVG rendering
# ====================================================================================================================


def _svg_chart(metric: str, points: dict[str, list[dict[str, float | int]]]) -> str:
    width, height = 760, 460
    left, right, top, bottom = 72, 64, 48, 78
    plot_width = width - left - right
    plot_height = height - top - bottom
    all_points = points["baseline"] + points["csf"]
    min_x = min(float(point["bpp"]) for point in all_points)
    max_x = max(float(point["bpp"]) for point in all_points)
    min_y = min(float(point["quality"]) for point in all_points)
    max_y = max(float(point["quality"]) for point in all_points)
    if min_x == max_x:
        min_x -= 0.01
        max_x += 0.01
    if min_y == max_y:
        min_y -= 0.01
        max_y += 0.01

    x_pad = (max_x - min_x) * 0.04
    y_pad = (max_y - min_y) * 0.08
    min_x = max(0.0, min_x - x_pad)
    max_x += x_pad
    min_y -= y_pad
    max_y += y_pad

    def sx(value: float) -> float:
        return left + (value - min_x) * plot_width / (max_x - min_x)

    def sy(value: float) -> float:
        return top + (max_y - value) * plot_height / (max_y - min_y)

    def clamp_label(x: float, y: float, mode: str) -> tuple[float, float]:
        x = min(max(x + 6, left + 4), width - right - 34)
        offset = -8 if mode == "baseline" else 16
        y = min(max(y + offset, top + 14), height - bottom - 4)
        return x, y

    colors = {"baseline": "#2563eb", "csf": "#dc2626"}
    styles = {
        "baseline": {"width": "3.4", "dash": "", "marker": "circle"},
        "csf": {"width": "2.6", "dash": ' stroke-dasharray="7 5"', "marker": "diamond"},
    }
    y_label = METRIC_CHART_LABELS.get(metric, metric)
    title = y_label.replace(", dB", "")
    lines = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        f'<text x="{left}" y="26" font-family="Arial" font-size="18" font-weight="700">{html.escape(title)} average RD curve</text>',
    ]

    for value in _linear_ticks(min_x, max_x):
        x = sx(value)
        lines.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{height - bottom}" stroke="#e5e7eb" stroke-width="1"/>')
    for value in _linear_ticks(min_y, max_y):
        y = sy(value)
        lines.append(f'<line x1="{left}" y1="{y:.1f}" x2="{width - right}" y2="{y:.1f}" stroke="#e5e7eb" stroke-width="1"/>')

    lines.extend(
        [
            f'<line x1="{left}" y1="{height - bottom}" x2="{width - right}" y2="{height - bottom}" stroke="#111827"/>',
            f'<line x1="{left}" y1="{top}" x2="{left}" y2="{height - bottom}" stroke="#111827"/>',
            f'<text x="{width // 2 - 12}" y="{height - 28}" font-family="Arial" font-size="13">bpp</text>',
            f'<text x="16" y="{height // 2 + 45}" font-family="Arial" font-size="13" transform="rotate(-90 16,{height // 2 + 45})">{html.escape(y_label)}</text>',
        ]
    )

    for value in _linear_ticks(min_x, max_x):
        x = sx(value)
        lines.extend(
            [
                f'<line x1="{x:.1f}" y1="{height - bottom}" x2="{x:.1f}" y2="{height - bottom + 5}" stroke="#111827"/>',
                f'<text x="{x:.1f}" y="{height - bottom + 20}" font-family="Arial" font-size="11" fill="#374151" text-anchor="middle">{_format_bpp_tick(value)}</text>',
            ]
        )
    for value in _linear_ticks(min_y, max_y):
        y = sy(value)
        lines.extend(
            [
                f'<line x1="{left - 5}" y1="{y:.1f}" x2="{left}" y2="{y:.1f}" stroke="#111827"/>',
                f'<text x="{left - 9}" y="{y + 4:.1f}" font-family="Arial" font-size="11" fill="#374151" text-anchor="end">{_format_quality_tick(value)}</text>',
            ]
        )

    for mode in ("baseline", "csf"):
        mode_points = points.get(mode, [])
        if not mode_points:
            continue
        color = colors[mode]
        style = styles[mode]
        coords = [(sx(float(point["bpp"])), sy(float(point["quality"]))) for point in mode_points]
        lines.append(
            '<polyline fill="none" stroke="{color}" stroke-width="{width}" stroke-linecap="round" '
            'stroke-linejoin="round"{dash} points="{points}"/>'.format(
                color=color,
                width=style["width"],
                dash=style["dash"],
                points=" ".join(f"{x:.1f},{y:.1f}" for x, y in coords),
            )
        )
        for point, (x, y) in zip(mode_points, coords):
            lines.append(_marker_svg(x, y, color, str(style["marker"])))
            label_x, label_y = clamp_label(x, y, mode)
            lines.append(f'<text x="{label_x:.1f}" y="{label_y:.1f}" font-family="Arial" font-size="11" fill="#374151">QP{int(point["qp"])}</text>')

    lines.extend(
        [
            '<rect x="88" y="54" width="126" height="48" rx="4" fill="#ffffff" stroke="#e5e7eb"/>',
            f'<line x1="100" y1="72" x2="130" y2="72" stroke="{colors["baseline"]}" stroke-width="{styles["baseline"]["width"]}" stroke-linecap="round"/>',
            _marker_svg(115, 72, colors["baseline"], str(styles["baseline"]["marker"])),
            '<text x="140" y="76" font-family="Arial" font-size="12">baseline</text>',
            f'<line x1="100" y1="90" x2="130" y2="90" stroke="{colors["csf"]}" stroke-width="{styles["csf"]["width"]}" stroke-linecap="round" stroke-dasharray="7 5"/>',
            _marker_svg(115, 90, colors["csf"], str(styles["csf"]["marker"])),
            '<text x="140" y="94" font-family="Arial" font-size="12">csf</text>',
            "</svg>",
        ]
    )
    return "\n".join(lines)


def _svg_qp_chart(image: str, metric: str, points: dict[str, list[dict[str, float | int]]]) -> str:
    width, height = 760, 420
    left, right, top, bottom = 72, 64, 48, 70
    plot_width = width - left - right
    plot_height = height - top - bottom
    all_points = points["baseline"] + points["csf"]
    min_x = min(float(point["qp"]) for point in all_points)
    max_x = max(float(point["qp"]) for point in all_points)
    min_y = min(float(point["quality"]) for point in all_points)
    max_y = max(float(point["quality"]) for point in all_points)
    if min_x == max_x:
        min_x -= 1
        max_x += 1
    if min_y == max_y:
        min_y -= 0.01
        max_y += 0.01

    y_pad = (max_y - min_y) * 0.08
    min_y -= y_pad
    max_y += y_pad

    def sx(value: float) -> float:
        return left + (value - min_x) * plot_width / (max_x - min_x)

    def sy(value: float) -> float:
        return top + (max_y - value) * plot_height / (max_y - min_y)

    colors = {"baseline": "#2563eb", "csf": "#dc2626"}
    styles = {
        "baseline": {"width": "3.2", "dash": "", "marker": "circle"},
        "csf": {"width": "2.6", "dash": ' stroke-dasharray="7 5"', "marker": "diamond"},
    }
    y_label = METRIC_CHART_LABELS.get(metric, metric)
    title = f"{image}: {y_label.replace(', dB', '')} vs QP"
    lines = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        f'<text x="{left}" y="26" font-family="Arial" font-size="18" font-weight="700">{html.escape(title)}</text>',
    ]

    qps = sorted({int(point["qp"]) for point in all_points})
    for qp in qps:
        x = sx(qp)
        lines.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{height - bottom}" stroke="#e5e7eb" stroke-width="1"/>')
    for value in _linear_ticks(min_y, max_y):
        y = sy(value)
        lines.append(f'<line x1="{left}" y1="{y:.1f}" x2="{width - right}" y2="{y:.1f}" stroke="#e5e7eb" stroke-width="1"/>')

    lines.extend(
        [
            f'<line x1="{left}" y1="{height - bottom}" x2="{width - right}" y2="{height - bottom}" stroke="#111827"/>',
            f'<line x1="{left}" y1="{top}" x2="{left}" y2="{height - bottom}" stroke="#111827"/>',
            f'<text x="{width // 2 - 12}" y="{height - 26}" font-family="Arial" font-size="13">QP</text>',
            f'<text x="16" y="{height // 2 + 45}" font-family="Arial" font-size="13" transform="rotate(-90 16,{height // 2 + 45})">{html.escape(y_label)}</text>',
        ]
    )

    for qp in qps:
        x = sx(qp)
        lines.extend(
            [
                f'<line x1="{x:.1f}" y1="{height - bottom}" x2="{x:.1f}" y2="{height - bottom + 5}" stroke="#111827"/>',
                f'<text x="{x:.1f}" y="{height - bottom + 20}" font-family="Arial" font-size="11" fill="#374151" text-anchor="middle">{qp}</text>',
            ]
        )
    for value in _linear_ticks(min_y, max_y):
        y = sy(value)
        lines.extend(
            [
                f'<line x1="{left - 5}" y1="{y:.1f}" x2="{left}" y2="{y:.1f}" stroke="#111827"/>',
                f'<text x="{left - 9}" y="{y + 4:.1f}" font-family="Arial" font-size="11" fill="#374151" text-anchor="end">{_format_quality_tick(value)}</text>',
            ]
        )

    for mode in ("baseline", "csf"):
        mode_points = points.get(mode, [])
        if not mode_points:
            continue
        color = colors[mode]
        style = styles[mode]
        coords = [(sx(float(point["qp"])), sy(float(point["quality"]))) for point in mode_points]
        lines.append(
            '<polyline fill="none" stroke="{color}" stroke-width="{width}" stroke-linecap="round" '
            'stroke-linejoin="round"{dash} points="{points}"/>'.format(
                color=color,
                width=style["width"],
                dash=style["dash"],
                points=" ".join(f"{x:.1f},{y:.1f}" for x, y in coords),
            )
        )
        for x, y in coords:
            lines.append(_marker_svg(x, y, color, str(style["marker"])))

    lines.extend(
        [
            f'<rect x="{width - right - 132}" y="{top + 6}" width="126" height="48" rx="4" fill="#ffffff" stroke="#e5e7eb"/>',
            f'<line x1="{width - right - 120}" y1="{top + 24}" x2="{width - right - 90}" y2="{top + 24}" stroke="{colors["baseline"]}" stroke-width="{styles["baseline"]["width"]}" stroke-linecap="round"/>',
            _marker_svg(width - right - 105, top + 24, colors["baseline"], str(styles["baseline"]["marker"])),
            f'<text x="{width - right - 80}" y="{top + 28}" font-family="Arial" font-size="12">baseline</text>',
            f'<line x1="{width - right - 120}" y1="{top + 42}" x2="{width - right - 90}" y2="{top + 42}" stroke="{colors["csf"]}" stroke-width="{styles["csf"]["width"]}" stroke-linecap="round" stroke-dasharray="7 5"/>',
            _marker_svg(width - right - 105, top + 42, colors["csf"], str(styles["csf"]["marker"])),
            f'<text x="{width - right - 80}" y="{top + 46}" font-family="Arial" font-size="12">csf</text>',
            "</svg>",
        ]
    )
    return "\n".join(lines)


def _marker_svg(x: float, y: float, color: str, marker: str) -> str:
    if marker == "diamond":
        return (
            f'<rect x="{x - 4:.1f}" y="{y - 4:.1f}" width="8" height="8" '
            f'transform="rotate(45 {x:.1f} {y:.1f})" fill="#ffffff" stroke="{color}" stroke-width="2"/>'
        )
    return f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4.5" fill="#ffffff" stroke="{color}" stroke-width="2"/>'


def _linear_ticks(min_value: float, max_value: float, count: int = 5) -> list[float]:
    if count <= 1 or min_value == max_value:
        return [min_value]
    step = (max_value - min_value) / (count - 1)
    return [min_value + step * index for index in range(count)]


def _format_bpp_tick(value: float) -> str:
    return f"{value:.3f}" if value < 1 else f"{value:.2f}"


def _format_quality_tick(value: float) -> str:
    return f"{value:.1f}" if abs(value) >= 10 else f"{value:.4f}".rstrip("0").rstrip(".")


def _safe_name(value: str) -> str:
    return "".join(char if char.isalnum() or char in "-_" else "_" for char in value)


class ImageBenchmarkReportBuilder:
    """Builds CSV summaries and SVG charts from an existing image_metrics.csv file."""

    def __init__(self, metrics_csv: Path, output: Path, write_xlsx_output: bool = False) -> None:
        self.metrics_csv = metrics_csv
        self.output = output
        self.write_xlsx_output = write_xlsx_output

    def build(self) -> None:
        rows = read_rows(self.metrics_csv)
        same_qp = same_qp_deltas(rows)
        equal_bpp = equal_bpp_summary(rows)
        bd_rows = bd_rate_summary(rows)
        bd_summary = aggregate_bd_rate(bd_rows) if bd_rows else []
        write_csv(self.output / "same_qp_deltas.csv", same_qp)
        write_csv(self.output / "per_image_summary.csv", per_image_summary(same_qp))
        write_csv(self.output / "equal_bpp_summary.csv", equal_bpp)
        write_csv(self.output / "bd_rate_by_image.csv", bd_rows)
        write_csv(self.output / "bd_rate_summary.csv", bd_summary)
        if same_qp:
            aggregate_summary(same_qp, self.output / "same_qp_summary.csv")
        if equal_bpp:
            aggregate_summary(equal_bpp, self.output / "equal_bpp_metric_summary.csv")
        if self.write_xlsx_output:
            write_xlsx(self.output / "results.xlsx", rows, same_qp, bd_summary)
        render_metric_charts(rows, self.output / "charts")
        render_per_image_qp_charts(rows, self.output / "qp_charts")


def main() -> int:
    parser = argparse.ArgumentParser(description="Create summary CSVs and SVG charts for image CSF benchmark metrics.")
    parser.add_argument("metrics_csv", type=Path)
    parser.add_argument("--output", type=Path, default=Path("docs/image_benchmark"))
    parser.add_argument("--xlsx", action="store_true", help="Write results.xlsx with raw metrics, same-QP summary, and BD-Rate sheets.")
    args = parser.parse_args()

    ImageBenchmarkReportBuilder(args.metrics_csv, args.output, write_xlsx_output=args.xlsx).build()
    logger.info("Wrote image benchmark report files to %s", args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
